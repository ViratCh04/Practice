import openpyxl

wb = openpyxl.Workbook()
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value == None)
sheet['A1'] = 'Hello'
sheet['A2'] = 53

import os
os.chdir('C:\\users\\USER\\Documents')

wb.save('hiha.xlsx')
sheet2 = wb.create_sheet()
print(wb.get_sheet_names())

print(sheet2.title)
sheet2.title = 'NewSheetName'
print(wb.get_sheet_names())

wb.save('hiha2.xlsx')
wb.create_sheet(index=0, title='OtherSheet')
wb.save('hiha2.xlsx')
