import openpyxl
import os

os.chdir('C:\\users\\USER\\Downloads')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

print(workbook.get_sheet_names())
print(sheet['B1'].value)
# or you can also store it in a variable
cell = sheet['A1']
print(cell.value)

print(sheet.cell(row=1, column=3))

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
    